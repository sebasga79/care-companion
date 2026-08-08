"""`DatasetCaseAdapter` — `ChallengeCasePort` real sobre el dataset oficial
del reto (docs/auditoria-kit-oficial-2026-08-07.md §4.2/§9.2: no es Delta
Share, son 4 `.xlsx` — descargados por `scripts/fetch_dataset.py` a
`DATASET_DIR`, nunca commiteados al repo, ver `.gitignore`).

Lee `trayectorias_postop_silver.xlsx` (una fila = un hito histórico:
paciente × día postoperatorio), `perfiles_clinicos_pacientes_silver_contest.xlsx`
(procedimiento/edad/género/comorbilidades) y `perfiles_pacientes_co.xlsx`
(demografía colombiana). `dataset_final.xlsx` (las 3.991 conversaciones
guionizadas) NO se usa aquí — no hace falta para poblar el selector de
casos de `/call`; es material para un futuro arnés de evaluación
automatizada (fuera de alcance de este adapter).

Join real del dataset (no inventado, confirmado inspeccionando los
archivos con `openpyxl`): `paciente_id` conecta los tres archivos;
`caso_id = "caso_" + trayectoria_id`. La superficie de llamada agrupa esos
160 episodios en 40 entidades de paciente; los IDs originales permanecen
disponibles internamente para trazabilidad y pruebas del kit."""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.ports.challenge_case import (
    CaseFilters,
    CaseSummary,
    ChallengeCase,
    ChallengeCasePort,
    HistoricalFollowup,
    ReferenceTrajectory,
)

logger = logging.getLogger("care_companion.dataset")

TRAYECTORIAS_FILE = "trayectorias_postop_silver.xlsx"
PERFILES_CLINICOS_FILE = "perfiles_clinicos_pacientes_silver_contest.xlsx"
PERFILES_DEMOGRAFICOS_FILE = "perfiles_pacientes_co.xlsx"

REQUIRED_FILES = (TRAYECTORIAS_FILE, PERFILES_CLINICOS_FILE, PERFILES_DEMOGRAFICOS_FILE)


class DatasetFilesMissingError(Exception):
    """Falta alguno de los `.xlsx` requeridos en `dataset_dir`. El caller
    (`main.py`) decide qué hacer — hoy, caer a `FixtureCaseAdapter` con un
    warning explícito en el log, nunca fingir que el dataset real está
    cargado cuando no lo está (spec.md §11.2)."""

    def __init__(self, dataset_dir: Path, missing: list[str]) -> None:
        self.dataset_dir = dataset_dir
        self.missing = missing
        super().__init__(
            f"Faltan archivos del dataset en {dataset_dir}: {', '.join(missing)} "
            "— correr `uv run python scripts/fetch_dataset.py`."
        )


def _read_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(cell) for cell in next(rows)]
        return [dict(zip(header, row, strict=True)) for row in rows if any(row)]
    finally:
        wb.close()


def _parse_json_cell(value: Any) -> list[str]:
    """`comorbilidades` viene como una lista JSON dentro de una celda de
    texto (README oficial del kit, sección "Antes de que empieces"). Si el
    valor no es JSON válido, se registra vacío en vez de romper la carga
    completa por una fila con datos sucios — un caso con metadata
    incompleta es preferible a que el selector de casos entero no cargue."""
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except (json.JSONDecodeError, TypeError):
        logger.warning("dataset_comorbilidades_json_invalido valor=%r", value)
        return []
    return [str(item) for item in parsed] if isinstance(parsed, list) else []


@dataclass(frozen=True)
class _PatientClinicalProfile:
    procedure: str
    procedure_category: str
    surgery_date: date | None
    age: int | None
    gender: str | None
    comorbidities: list[str]


@dataclass(frozen=True)
class _PatientDemographics:
    display_name: str
    city: str | None
    department: str | None


def _load_clinical_profiles(dataset_dir: Path) -> dict[str, _PatientClinicalProfile]:
    profiles: dict[str, _PatientClinicalProfile] = {}
    for row in _read_rows(dataset_dir / PERFILES_CLINICOS_FILE):
        paciente_id = row["paciente_id"]
        raw_surgery_date = row.get("fecha_cirugia")
        surgery_date = (
            raw_surgery_date.date()
            if isinstance(raw_surgery_date, datetime)
            else raw_surgery_date
            if isinstance(raw_surgery_date, date)
            else date.fromisoformat(raw_surgery_date)
            if isinstance(raw_surgery_date, str) and raw_surgery_date
            else None
        )
        profiles[paciente_id] = _PatientClinicalProfile(
            procedure=row.get("procedimiento") or "",
            procedure_category=row.get("modulo_synthea") or "",
            surgery_date=surgery_date,
            age=row.get("edad"),
            gender=row.get("genero"),
            comorbidities=_parse_json_cell(row.get("comorbilidades")),
        )
    return profiles


def _load_demographics(dataset_dir: Path) -> dict[str, _PatientDemographics]:
    demographics: dict[str, _PatientDemographics] = {}
    for row in _read_rows(dataset_dir / PERFILES_DEMOGRAFICOS_FILE):
        paciente_id = row["paciente_id"]
        demographics[paciente_id] = _PatientDemographics(
            display_name=row.get("nombre_completo") or paciente_id,
            city=row.get("ciudad"),
            department=row.get("departamento"),
        )
    return demographics


def _build_cases(
    dataset_dir: Path,
) -> tuple[dict[str, ChallengeCase], dict[str, ChallengeCase]]:
    clinical = _load_clinical_profiles(dataset_dir)
    demographics = _load_demographics(dataset_dir)

    episode_cases: dict[str, ChallengeCase] = {}
    history_by_patient: dict[str, list[HistoricalFollowup]] = {}
    for row in _read_rows(dataset_dir / TRAYECTORIAS_FILE):
        trayectoria_id = row["trayectoria_id"]
        paciente_id = row["paciente_id"]
        case_id = f"caso_{trayectoria_id}"
        dia_postop = int(row["dia_postop"])

        profile = clinical.get(paciente_id)
        demo = demographics.get(paciente_id)
        if profile is None or demo is None:
            # Fila de trayectoria sin perfil clínico/demográfico asociado —
            # dato incompleto, se omite ese caso en vez de fabricar
            # valores por defecto que parecerían reales.
            logger.warning(
                "dataset_trayectoria_sin_perfil paciente_id=%s trayectoria_id=%s",
                paciente_id,
                trayectoria_id,
            )
            continue

        historical_followup = HistoricalFollowup(
            trajectory_id=str(trayectoria_id),
            day=dia_postop,
            archetype=row.get("arquetipo_trayectoria") or "",
            pain_nrs=int(row["dolor_nrs"]) if row.get("dolor_nrs") is not None else 0,
            temperature_c=float(row["fiebre_c"]) if row.get("fiebre_c") is not None else 0.0,
            mobility=row.get("movilidad") or "",
            wound=row.get("herida") or "",
            appetite=row.get("apetito") or "",
            sleep=row.get("sueno") or "",
        )
        history_by_patient.setdefault(paciente_id, []).append(historical_followup)

        episode_cases[case_id] = ChallengeCase(
            case_id=case_id,
            patient_id=paciente_id,
            patient_display_name=demo.display_name,
            procedure=profile.procedure,
            procedure_category=profile.procedure_category,
            phase=f"post_discharge_day_{dia_postop}",
            days_since_procedure=dia_postop,
            caregiver_role="paciente o acompañante",
            notes=(
                "Caso del dataset oficial del reto (sintético, adaptado a "
                "Colombia) — no corresponde a una persona real."
            ),
            age=profile.age,
            gender=profile.gender,
            comorbidities=profile.comorbidities,
            city=demo.city,
            department=demo.department,
            surgery_date=profile.surgery_date,
            reference_trajectory=ReferenceTrajectory(
                arquetipo=historical_followup.archetype,
                dolor_nrs=historical_followup.pain_nrs,
                fiebre_c=historical_followup.temperature_c,
                movilidad=historical_followup.mobility,
                herida=historical_followup.wound,
                apetito=historical_followup.appetite,
                sueno=historical_followup.sleep,
            ),
        )

    patient_cases: dict[str, ChallengeCase] = {}
    for paciente_id, history in history_by_patient.items():
        profile = clinical[paciente_id]
        demo = demographics[paciente_id]
        ordered_history = sorted(history, key=lambda item: item.day)
        latest_day = ordered_history[-1].day
        patient_case_id = f"paciente_{paciente_id}"
        patient_cases[patient_case_id] = ChallengeCase(
            case_id=patient_case_id,
            patient_id=paciente_id,
            patient_display_name=demo.display_name,
            procedure=profile.procedure,
            procedure_category=profile.procedure_category,
            phase="longitudinal_follow_up",
            days_since_procedure=latest_day,
            caregiver_role="paciente o acompañante",
            notes=(
                "Entidad longitudinal del dataset oficial: consolida los "
                "seguimientos históricos disponibles antes de la nueva llamada."
            ),
            age=profile.age,
            gender=profile.gender,
            comorbidities=profile.comorbidities,
            city=demo.city,
            department=demo.department,
            surgery_date=profile.surgery_date,
            historical_followups=ordered_history,
        )
    return patient_cases, episode_cases


def check_dataset_files_present(dataset_dir: Path) -> list[str]:
    """Devuelve la lista de archivos requeridos que faltan (vacía si están
    todos) — el caller decide si eso es fatal o si cae a fixtures."""
    return [name for name in REQUIRED_FILES if not (dataset_dir / name).is_file()]


class DatasetCaseAdapter(ChallengeCasePort):
    """Carga los 3 `.xlsx` una sola vez en memoria al construirse (40
    pacientes, 160 episodios históricos; trivial en tamaño) — no hay necesidad de releer disco por
    request; los archivos no cambian en caliente durante una sesión del
    reto."""

    def __init__(self, dataset_dir: str | Path) -> None:
        dataset_dir = Path(dataset_dir)
        missing = check_dataset_files_present(dataset_dir)
        if missing:
            raise DatasetFilesMissingError(dataset_dir, missing)
        self._cases, self._episode_cases = _build_cases(dataset_dir)
        logger.info(
            "dataset_case_adapter_loaded patient_count=%d episode_count=%d",
            len(self._cases),
            len(self._episode_cases),
        )

    async def list_cases(self, filters: CaseFilters) -> list[CaseSummary]:
        cases = list(self._cases.values())
        if filters.procedure:
            needle = filters.procedure.strip().lower()
            cases = [
                c
                for c in cases
                if c.procedure.lower() == needle or c.procedure_category.lower() == needle
            ]
        cases = cases[: filters.limit]
        return [
            CaseSummary(
                case_id=c.case_id,
                patient_id=c.patient_id,
                patient_display_name=c.patient_display_name,
                procedure=c.procedure,
                procedure_category=c.procedure_category,
                phase=c.phase,
                days_since_procedure=c.days_since_procedure,
                surgery_date=c.surgery_date,
                followup_days=[item.day for item in c.historical_followups],
                historical_followups=list(c.historical_followups),
            )
            for c in cases
        ]

    async def get_case(self, case_id: str) -> ChallengeCase | None:
        return self._cases.get(case_id) or self._episode_cases.get(case_id)


__all__ = [
    "DatasetCaseAdapter",
    "DatasetFilesMissingError",
    "check_dataset_files_present",
]
