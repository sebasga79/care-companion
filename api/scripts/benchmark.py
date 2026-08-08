#!/usr/bin/env python3
"""Benchmark automatizado contra las conversaciones REALES del dataset del
reto (`dataset_final.xlsx`), no contra casos inventados.

Qué mide, y por qué cada cosa:

- **Escalamiento vs `label_ground_truth`.** El dataset trae la criticidad de
  referencia de cada caso (`verde`/`amarillo`/`rojo`). Reproducimos los
  turnos reales del paciente y comparamos con lo que decide el sistema. El
  número que manda es el **falso negativo** (un caso `rojo` que no escaló):
  la rúbrica lo llama la falla catastrófica y puede anular por sí solo el
  criterio de 20 pts de decisión y escalamiento.
- **Latencia P50/P95** por turno conversacional — obligatoria en el README.
- **Tokens de entrada/salida** por turno y por llamada, invocaciones al
  modelo por turno y consultas al RAG por llamada — también obligatorias.

Las dos capas del dataset se pueden correr por separado: `capa1_limpia`
(el paciente responde lo que se le pregunta) y `capa2_ruidosa` (respuestas
evasivas, información faltante, interrupciones de terceros). Comparar ambas
muestra cuánto se degrada el sistema con habla realista.

Uso:
    uv run python scripts/benchmark.py --limit 10
    uv run python scripts/benchmark.py --capa capa2_ruidosa --limit 10
    uv run python scripts/benchmark.py --limit 40 --out docs/benchmarks/run.json

Corre con el proveedor configurado en `.env`. Con `LLM_PROVIDER=fake` es
determinista e instantáneo (útil para verificar el arnés); con `groq` u
`ollama` mide el sistema real.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import statistics
import sys
import time
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import get_settings
from app.main import create_app

CONVERSATIONS_FILE = "dataset_final.xlsx"

# Estados en los que la sesión ya no acepta más turnos.
TERMINAL_STATES = {"summarizing", "closed", "fail_safe", "escalated"}

# Un caso `rojo` DEBE terminar escalado; un `verde` NO debería. `amarillo`
# es deliberadamente ambiguo —el propio kit lo describe como zona gris— así
# que se reporta aparte y no cuenta como acierto ni como error.
LABEL_SHOULD_ESCALATE = {"rojo": True, "verde": False}


def load_conversations(dataset_dir: Path, capa: str) -> dict[str, dict[str, Any]]:
    """Agrupa los turnos del PACIENTE por caso. Los turnos del agente se
    descartan: aquí el agente es el sistema bajo prueba."""
    path = dataset_dir / CONVERSATIONS_FILE
    if not path.is_file():
        print(f"ERROR: falta {path} — correr `scripts/fetch_dataset.py`.", file=sys.stderr)
        raise SystemExit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c) for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}
        cases: dict[str, dict[str, Any]] = {}
        for row in rows:
            if not row or row[idx["capa"]] != capa:
                continue
            case_id = row[idx["caso_id"]]
            entry = cases.setdefault(
                case_id,
                {
                    "case_id": case_id,
                    "patient_id": row[idx["paciente_id"]],
                    "label": row[idx["label_ground_truth"]],
                    "day": row[idx["dia_postop"]],
                    "style": row[idx["estilo_paciente"]],
                    "turns": [],
                },
            )
            if row[idx["hablante"]] == "paciente" and row[idx["texto"]]:
                entry["turns"].append((row[idx["turno_idx"]], str(row[idx["texto"]])))
        for entry in cases.values():
            entry["turns"] = [text for _, text in sorted(entry["turns"])]
        return cases
    finally:
        wb.close()


def _percentile(values: list[float], fraction: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    index = max(0, min(len(ordered) - 1, round(fraction * (len(ordered) - 1))))
    return ordered[index]


class TokenBudget:
    """Throttle por tokens/minuto: espera lo justo para que la ventana
    deslizante vuelva a tener presupuesto.

    Es la forma simple de respetar la cuota: en vez de chocar contra el 429 y
    reaccionar, se lleva la cuenta de lo consumido en los últimos 60 s y se
    espera antes de pedir más. Determinista y sin sorpresas."""

    def __init__(self, tokens_per_minute: int, window_seconds: float = 60.0) -> None:
        self._limit = tokens_per_minute
        self._window = window_seconds
        self._spent: list[tuple[float, int]] = []

    def _prune(self, now: float) -> None:
        self._spent = [(ts, n) for ts, n in self._spent if now - ts < self._window]

    async def reserve(self, estimated_tokens: int) -> float:
        """Espera hasta que quepan `estimated_tokens`. Devuelve lo esperado."""
        waited = 0.0
        while True:
            now = time.monotonic()
            self._prune(now)
            used = sum(n for _, n in self._spent)
            if used + estimated_tokens <= self._limit or not self._spent:
                return waited
            oldest = min(ts for ts, _ in self._spent)
            sleep_for = max(0.5, self._window - (now - oldest) + 0.5)
            await asyncio.sleep(sleep_for)
            waited += sleep_for

    def record(self, tokens: int) -> None:
        self._spent.append((time.monotonic(), tokens))


async def run_case(
    app: Any, patient_case_id: str, turns: list[str], budget: TokenBudget, est_per_turn: int
) -> dict[str, Any]:
    """Reproduce una conversación llamando al orquestador DIRECTAMENTE.

    Antes esto pasaba por `TestClient` + WebSocket, y ahí es donde el
    benchmark se colgaba: una espera larga dentro del handler bloquea el
    ciclo de eventos que el WebSocket de pruebas necesita para entregar los
    mensajes. El orquestador es una corrutina normal — llamarla directo
    elimina el problema de raíz y permite throttlear con precisión.

    Se pierde la verificación del contrato de envelopes del WebSocket, que
    ya cubren `tests/test_ws.py` y `tests/test_gates.py`.
    """
    from app.repositories.knowledge import get_current_knowledge_version

    settings = app.state.settings
    record = app.state.session_repo.create(
        case_id=patient_case_id,
        state="created",
        knowledge_version=get_current_knowledge_version(settings.database_path),
    )
    session_id = record["id"]
    orchestrator = app.state.call_cycle_orchestrator

    latencies: list[float] = []
    escalated = False
    decision_level: str | None = None
    state = "created"
    turns_sent = 0

    for text in turns:
        waited = await budget.reserve(est_per_turn)
        started = time.perf_counter()
        result = await orchestrator.handle_turn(session_id, text)
        elapsed_ms = (time.perf_counter() - started) * 1000
        latencies.append(elapsed_ms)
        turns_sent += 1

        state = result.state.value
        decision_level = result.decision_level.value
        escalated = escalated or bool(result.escalated)
        budget.record(est_per_turn)
        if waited:
            logger_note = f" (esperó {waited:.0f}s de cuota)"
        else:
            logger_note = ""
        print(f"      turno {turns_sent}: {elapsed_ms:.0f} ms{logger_note}", flush=True)

        if state in TERMINAL_STATES:
            break

    return {
        "session_id": session_id,
        "state": state,
        "decision_level": decision_level,
        "escalated": escalated,
        "turns_sent": turns_sent,
        "turns_available": len(turns),
        "latencies_ms": latencies,
    }


def collect_usage(app: Any, session_id: str) -> dict[str, int]:
    """Uso real del turno leído de los eventos persistidos, no estimado."""
    tokens_in = tokens_out = llm_calls = rag_queries = 0
    for event in app.state.event_repo.list_for_session(session_id):
        payload = event.get("payload")
        data = json.loads(payload) if isinstance(payload, str) else (payload or {})
        if event["event_type"].startswith("agent.") and event["event_type"].endswith(
            ".completed"
        ):
            tokens_in += int(data.get("input_tokens", 0) or 0)
            tokens_out += int(data.get("output_tokens", 0) or 0)
            llm_calls += 1
        elif event["event_type"] == "rag.retrieval.completed":
            rag_queries += 1
    return {
        "input_tokens": tokens_in,
        "output_tokens": tokens_out,
        "llm_calls": llm_calls,
        "rag_queries": rag_queries,
    }


def summarize(results: list[dict[str, Any]], settings: Any, capa: str) -> dict[str, Any]:
    latencies = [ms for r in results for ms in r.get("latencies_ms", [])]
    confusion: dict[str, dict[str, int]] = defaultdict(lambda: {"escalated": 0, "not": 0})
    false_negatives: list[dict[str, Any]] = []
    false_positives: list[dict[str, Any]] = []

    tokens_in = sum(r.get("usage", {}).get("input_tokens", 0) for r in results)
    tokens_out = sum(r.get("usage", {}).get("output_tokens", 0) for r in results)
    llm_calls = sum(r.get("usage", {}).get("llm_calls", 0) for r in results)
    rag_queries = sum(r.get("usage", {}).get("rag_queries", 0) for r in results)
    total_turns = sum(r.get("turns_sent", 0) for r in results)

    for result in results:
        label = result["label"]
        confusion[label]["escalated" if result["escalated"] else "not"] += 1
        expected = LABEL_SHOULD_ESCALATE.get(label)
        if expected is True and not result["escalated"]:
            false_negatives.append(result)
        elif expected is False and result["escalated"]:
            false_positives.append(result)

    scored = [r for r in results if r["label"] in LABEL_SHOULD_ESCALATE]
    correct = len(scored) - len(false_negatives) - len(false_positives)
    red = [r for r in results if r["label"] == "rojo"]
    green = [r for r in results if r["label"] == "verde"]

    return {
        "generated_at": datetime.now(UTC).isoformat(),
        "capa": capa,
        "provider": settings.llm_provider.value,
        "model": settings.llm_model,
        "fallback": (
            settings.llm_fallback_provider.value if settings.llm_fallback_provider else None
        ),
        "embeddings": settings.embeddings_provider.value,
        "cases": len(results),
        "turns": total_turns,
        "decision": {
            "confusion": {k: dict(v) for k, v in confusion.items()},
            "false_negatives": len(false_negatives),
            "false_positives": len(false_positives),
            "false_negative_cases": [r["case_id"] for r in false_negatives],
            "false_positive_cases": [r["case_id"] for r in false_positives],
            "accuracy_scored": round(correct / len(scored), 4) if scored else None,
            "sensitivity_red": (
                round(sum(1 for r in red if r["escalated"]) / len(red), 4) if red else None
            ),
            "specificity_green": (
                round(sum(1 for r in green if not r["escalated"]) / len(green), 4)
                if green
                else None
            ),
        },
        "latency_ms": {
            "p50": round(_percentile(latencies, 0.50) or 0, 1),
            "p95": round(_percentile(latencies, 0.95) or 0, 1),
            "mean": round(statistics.fmean(latencies), 1) if latencies else None,
            "max": round(max(latencies), 1) if latencies else None,
            "samples": len(latencies),
        },
        "usage": {
            "input_tokens_total": tokens_in,
            "output_tokens_total": tokens_out,
            "input_tokens_per_turn": round(tokens_in / total_turns, 1) if total_turns else None,
            "output_tokens_per_turn": round(tokens_out / total_turns, 1) if total_turns else None,
            "tokens_per_call": round((tokens_in + tokens_out) / len(results), 1)
            if results
            else None,
            "llm_calls_per_turn": round(llm_calls / total_turns, 2) if total_turns else None,
            "rag_queries_per_call": round(rag_queries / len(results), 2) if results else None,
        },
    }



def _write_report(
    out: Path, results: list[dict[str, Any]], settings: Any, capa: str, *, partial: bool
) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    report = summarize(results, settings, capa)
    report["partial"] = partial
    out.write_text(
        json.dumps({"summary": report, "cases": results}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--limit", type=int, default=10, help="casos a correr (default: 10)")
    parser.add_argument(
        "--capa", default="capa1_limpia", choices=["capa1_limpia", "capa2_ruidosa"]
    )
    parser.add_argument("--max-turns", type=int, default=8, help="tope de turnos por caso")
    parser.add_argument("--out", default=None, help="ruta del JSON de resultados")
    parser.add_argument(
        "--tpm",
        type=int,
        default=6000,
        help="presupuesto de tokens por minuto del proveedor (Groq free: 6000)",
    )
    parser.add_argument(
        "--est-tokens",
        type=int,
        default=6500,
        help="tokens estimados por turno (3 agentes); se reserva antes de cada turno",
    )
    parser.add_argument(
        "--pause",
        type=float,
        default=0.0,
        help=(
            "segundos de espera entre casos. Con el nivel gratuito de Groq "
            "(6.000 tokens/minuto) una corrida seguida agota la cuota y el "
            "sistema degrada al resguardo, contaminando la medición de latencia."
        ),
    )
    args = parser.parse_args()

    settings = get_settings()
    conversations = load_conversations(Path(settings.dataset_dir), args.capa)

    # Muestra estratificada: los casos `rojo` son sólo 12 de 160 y son los
    # que más importan (falso negativo = falla catastrófica). Una muestra
    # puramente secuencial podría no incluir ninguno.
    by_label: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for case in conversations.values():
        by_label[case["label"]].append(case)
    for bucket in by_label.values():
        bucket.sort(key=lambda c: c["case_id"])

    selected: list[dict[str, Any]] = []
    order = ["rojo", "amarillo", "verde"]
    cursor = 0
    while len(selected) < args.limit and any(
        cursor < len(by_label[label]) for label in order
    ):
        for label in order:
            if cursor < len(by_label[label]) and len(selected) < args.limit:
                selected.append(by_label[label][cursor])
        cursor += 1

    # Medir contra el modelo declarado exige ESPERAR el rate limit, no
    # degradar al resguardo: si el 429 cae al modelo local, la latencia
    # medida es la del resguardo (20x más lenta) y el modelo evaluado ya no
    # es el que se declara para G3. Aquí sí se puede esperar — no hay un
    # paciente al teléfono.
    import os

    os.environ.setdefault("LLM_RATE_LIMIT_MAX_RETRIES", "3")
    os.environ.setdefault("LLM_RATE_LIMIT_MAX_WAIT_SECONDS", "20")
    settings = get_settings()

    app = create_app()
    budget = TokenBudget(args.tpm)
    print(
        f"provider={settings.llm_provider.value} model={settings.llm_model} "
        f"capa={args.capa} casos={len(selected)}\n",
        flush=True,
    )

    results: list[dict[str, Any]] = []
    for position, case in enumerate(selected, start=1):
        patient_case_id = f"paciente_{case['patient_id']}"
        turns = case["turns"][: args.max_turns]
        if not turns:
            continue
        print(f"  [{position}/{len(selected)}] {case['label']} {case['case_id']}", flush=True)
        outcome = asyncio.run(run_case(app, patient_case_id, turns, budget, args.est_tokens))
        outcome.update(
            {
                "case_id": case["case_id"],
                "label": case["label"],
                "day": case["day"],
                "style": case["style"],
                "usage": collect_usage(app, outcome["session_id"]),
            }
        )
        actual = outcome["usage"].get("input_tokens", 0) + outcome["usage"].get("output_tokens", 0)
        if actual and outcome["turns_sent"]:
            budget.record(max(0, actual - args.est_tokens * outcome["turns_sent"]))
        results.append(outcome)

        # Guardado incremental: una corrida contra el nivel gratuito puede
        # durar horas por las esperas de cuota. Perder todo por un fallo en
        # el último caso sería absurdo, y además permite mirar el avance sin
        # interrumpir el proceso.
        if args.out:
            _write_report(Path(args.out), results, settings, args.capa, partial=True)

        if args.pause:
            time.sleep(args.pause)
        expected = LABEL_SHOULD_ESCALATE.get(case["label"])
        mark = "·"
        if expected is True and not outcome["escalated"]:
            mark = "FALSO NEGATIVO"
        elif expected is False and outcome["escalated"]:
            mark = "falso positivo"
        print(
            f"      -> escalado={str(outcome['escalated']):5s} "
            f"{outcome['decision_level'] or '-':28s} {mark}",
            flush=True,
        )

    report = summarize(results, settings, args.capa)
    print("\n" + json.dumps(report, indent=2, ensure_ascii=False))

    if args.out:
        _write_report(Path(args.out), results, settings, args.capa, partial=False)
        print(f"\nResultados en {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
